"""
Art metadata editor — Flask server.
Run with: python art_editor.py
Opens a browser UI to browse images and edit Excel metadata.
"""

import os
import re
import json
import datetime
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory, send_file
import openpyxl

ART_DIR = Path(__file__).parent
EXCEL_PATH = ART_DIR / "art.xlsx"
COLUMNS = ["Fecha", "Nombre", "Inspiración", "Idea", "Lugar", "Técnica", "Colección", "Temática", "Poseedor"]

app = Flask(__name__)


def filename_to_date(fname):
    """Parse YYYY-MM-DD from a filename like '2024-06-04-.jpg' or '2026-04-7.jpg'."""
    stem = Path(fname).stem.rstrip("-")
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", stem)
    if m:
        try:
            return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def date_to_key(d):
    if d is None:
        return None
    if isinstance(d, datetime.datetime):
        d = d.date()
    return d.isoformat()


def load_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    return wb, ws, rows


def sync_excel():
    """Add any JPG not yet in the Excel as a new row with blank metadata."""
    wb, ws, rows = load_excel()

    existing_keys = set()
    for row in rows:
        d = row[0]
        if d:
            key = date_to_key(d)
            if key:
                existing_keys.add(key)

    jpgs = sorted(
        [f.name for f in ART_DIR.glob("*.jpg")],
        key=lambda x: filename_to_date(x) or datetime.date.min
    )

    added = 0
    for fname in jpgs:
        d = filename_to_date(fname)
        if d is None:
            continue
        key = d.isoformat()
        if key not in existing_keys:
            dt = datetime.datetime(d.year, d.month, d.day)
            new_row = [dt] + [None] * (len(COLUMNS) - 1)
            ws.append(new_row)
            existing_keys.add(key)
            added += 1

    if added:
        wb.save(EXCEL_PATH)
        print(f"Synced: added {added} missing JPG entries to Excel.")
    else:
        print("Excel already in sync — no new entries added.")

    return added


def get_all_records():
    """Return list of dicts sorted by date, each with filename and metadata."""
    _, _, rows = load_excel()

    jpgs_by_key = {}
    for fname in ART_DIR.glob("*.jpg"):
        d = filename_to_date(fname.name)
        if d:
            jpgs_by_key[d.isoformat()] = fname.name

    records = []
    for row in rows:
        d = row[0]
        if not d:
            continue
        key = date_to_key(d)
        if not key:
            continue
        rec = {"date": key, "filename": jpgs_by_key.get(key, ""), "meta": {}}
        for i, col in enumerate(COLUMNS):
            val = row[i]
            if isinstance(val, (datetime.datetime, datetime.date)):
                val = val.isoformat()
            rec["meta"][col] = val if val is not None else ""
        records.append(rec)

    records.sort(key=lambda r: r["date"])
    return records


def save_record(date_key, meta):
    """Update a single row in the Excel by date key."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        cell_date = row[0].value
        key = date_to_key(cell_date)
        if key == date_key:
            for i, col in enumerate(COLUMNS):
                val = meta.get(col, "")
                if i == 0:
                    continue  # never overwrite Fecha
                row[i].value = val if val != "" else None
            wb.save(EXCEL_PATH)
            return True
    return False


# ── Routes ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_file(ART_DIR / "art_editor.html")


@app.route("/img/<path:filename>")
def serve_image(filename):
    return send_from_directory(ART_DIR, filename)


@app.route("/api/records")
def api_records():
    return jsonify(get_all_records())


@app.route("/api/save", methods=["POST"])
def api_save():
    data = request.json
    ok = save_record(data["date"], data["meta"])
    return jsonify({"ok": ok})


# ── Main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Syncing Excel with JPG files...")
    sync_excel()
    print("Starting server at http://127.0.0.1:5000")
    import webbrowser
    webbrowser.open("http://127.0.0.1:5000")
    app.run(debug=False, port=5000)
